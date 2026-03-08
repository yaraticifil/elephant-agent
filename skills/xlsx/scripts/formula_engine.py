#!/usr/bin/env python3
"""
Verdent XLSX Formula Engine — recalculate, audit, and diagnose Excel formulas.

Modes:
  recalc   Recalculate all formulas via LibreOffice, then audit (default)
  audit    Scan for errors and gather formula stats without recalculating
  deps     Show formula dependency map (which cells reference which)

Usage:
  python recalc.py <file.xlsx>                    # recalc + audit
  python recalc.py <file.xlsx> --audit-only       # skip recalc, scan only
  python recalc.py <file.xlsx> --deps             # formula dependency map
  python recalc.py <file.xlsx> --timeout 60       # custom timeout
"""

import argparse, json, os, platform, re, subprocess, sys
from collections import defaultdict
from pathlib import Path

try:
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter
except ImportError:
    sys.exit("openpyxl is required: pip install openpyxl")

KNOWN_ERRORS = {"#VALUE!", "#DIV/0!", "#REF!", "#NAME?", "#NULL!", "#NUM!", "#N/A"}
CELL_REF_RE = re.compile(
    r"(?:(?P<sheet>'?[A-Za-z0-9_ ]+'?)!)?"
    r"(?P<col>\$?[A-Z]{1,3})(?P<row>\$?\d+)"
)


# ── LibreOffice macro bootstrap ──────────────────────────────────────

_MACRO_BODY = """\
<?xml version="1.0" encoding="UTF-8"?>
<!DOCTYPE script:module PUBLIC "-//OpenOffice.org//DTD OfficeDocument 1.0//EN" "module.dtd">
<script:module xmlns:script="http://openoffice.org/2000/script"
  script:name="Module1" script:language="StarBasic">
  Sub RecalcAndStore()
    Dim doc As Object
    doc = ThisComponent
    doc.calculateAll()
    doc.store()
    doc.close(True)
  End Sub
</script:module>"""


def _macro_dir():
    if platform.system() == "Darwin":
        return Path.home() / "Library/Application Support/LibreOffice/4/user/basic/Standard"
    return Path.home() / ".config/libreoffice/4/user/basic/Standard"


def ensure_macro():
    target = _macro_dir() / "Module1.xba"
    if target.exists() and "RecalcAndStore" in target.read_text():
        return True
    if not target.parent.exists():
        subprocess.run(
            ["soffice", "--headless", "--terminate_after_init"],
            capture_output=True, timeout=15,
        )
        target.parent.mkdir(parents=True, exist_ok=True)
    try:
        target.write_text(_MACRO_BODY)
        return True
    except OSError:
        return False


def _find_timeout_cmd():
    system = platform.system()
    if system == "Linux":
        return "timeout"
    if system == "Darwin":
        for candidate in ("gtimeout", "timeout"):
            try:
                subprocess.run([candidate, "--version"], capture_output=True, timeout=2)
                return candidate
            except (FileNotFoundError, subprocess.TimeoutExpired):
                continue
    return None


# ── Core: recalculate ────────────────────────────────────────────────

def run_recalc(filepath: Path, timeout: int = 30) -> dict:
    if not ensure_macro():
        return {"ok": False, "reason": "Cannot install LibreOffice macro"}

    abs_str = str(filepath.resolve())
    argv = [
        "soffice", "--headless", "--norestore",
        "vnd.sun.star.script:Standard.Module1.RecalcAndStore?language=Basic&location=application",
        abs_str,
    ]
    wrapper = _find_timeout_cmd()
    if wrapper:
        argv = [wrapper, str(timeout)] + argv

    proc = subprocess.run(argv, capture_output=True, text=True)
    if proc.returncode not in (0, 124):
        return {"ok": False, "reason": proc.stderr.strip() or "soffice exited with error"}
    return {"ok": True}


# ── Core: audit ──────────────────────────────────────────────────────

def audit_formulas(filepath: Path) -> dict:
    wb_vals = load_workbook(filepath, data_only=True, read_only=True)
    wb_fmls = load_workbook(filepath, data_only=False, read_only=True)

    sheet_reports = {}
    grand_errors = 0
    grand_formulas = 0

    for sname in wb_vals.sheetnames:
        ws_v = wb_vals[sname]
        ws_f = wb_fmls[sname]

        errors_by_type = defaultdict(list)
        formula_cells = []
        cell_count = 0

        for row_v, row_f in zip(ws_v.iter_rows(), ws_f.iter_rows()):
            for cv, cf in zip(row_v, row_f):
                cell_count += 1
                if isinstance(cf.value, str) and cf.value.startswith("="):
                    coord = f"{sname}!{get_column_letter(cf.column)}{cf.row}"
                    formula_cells.append(coord)
                    grand_formulas += 1
                if isinstance(cv.value, str):
                    for err_tag in KNOWN_ERRORS:
                        if err_tag in cv.value:
                            loc = f"{sname}!{get_column_letter(cv.column)}{cv.row}"
                            errors_by_type[err_tag].append(loc)
                            grand_errors += 1
                            break

        sheet_reports[sname] = {
            "cells": cell_count,
            "formulas": len(formula_cells),
            "errors": {k: {"count": len(v), "locations": v[:15]} for k, v in errors_by_type.items()},
        }

    wb_vals.close()
    wb_fmls.close()

    return {
        "status": "clean" if grand_errors == 0 else "has_errors",
        "sheets": sheet_reports,
        "totals": {
            "formulas": grand_formulas,
            "errors": grand_errors,
        },
    }


# ── Core: dependency map ─────────────────────────────────────────────

def dependency_map(filepath: Path, limit: int = 200) -> dict:
    wb = load_workbook(filepath, data_only=False, read_only=True)
    deps = {}
    count = 0

    for sname in wb.sheetnames:
        ws = wb[sname]
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    src = f"{sname}!{get_column_letter(cell.column)}{cell.row}"
                    refs = []
                    for m in CELL_REF_RE.finditer(cell.value):
                        sheet_part = m.group("sheet") or sname
                        sheet_part = sheet_part.strip("'")
                        refs.append(f"{sheet_part}!{m.group('col').replace('$','')}{m.group('row').replace('$','')}")
                    if refs:
                        deps[src] = {"formula": cell.value, "references": refs}
                        count += 1
                        if count >= limit:
                            wb.close()
                            return {"dependencies": deps, "truncated": True, "shown": count}

    wb.close()
    return {"dependencies": deps, "truncated": False, "shown": count}


# ── CLI ──────────────────────────────────────────────────────────────

def main():
    ap = argparse.ArgumentParser(description="Verdent XLSX Formula Engine")
    ap.add_argument("file", help="Excel file path")
    ap.add_argument("--audit-only", action="store_true", help="Scan without recalculating")
    ap.add_argument("--deps", action="store_true", help="Show formula dependency map")
    ap.add_argument("--timeout", type=int, default=30, help="LibreOffice timeout in seconds")
    args = ap.parse_args()

    fp = Path(args.file)
    if not fp.exists():
        sys.exit(json.dumps({"error": f"{args.file} not found"}))

    if args.deps:
        print(json.dumps(dependency_map(fp), indent=2))
        return

    if not args.audit_only:
        rc = run_recalc(fp, args.timeout)
        if not rc["ok"]:
            sys.exit(json.dumps({"error": rc["reason"]}))

    report = audit_formulas(fp)
    print(json.dumps(report, indent=2))


if __name__ == "__main__":
    main()
